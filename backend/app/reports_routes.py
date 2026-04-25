import base64
import math
import os
import logging
import traceback
from datetime import datetime
import requests
import io
from PIL import Image

from flask import Blueprint, jsonify, render_template
from google.cloud import firestore
from openpyxl.chart.label import DataLabelList
from openpyxl.chart import PieChart
from app.health import prepare_export_data

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

DB = firestore.Client()
reports_bp = Blueprint("reports_bp", __name__)
TILE_SIZE_MAP = 512
REPORT_TILE_URL = "https://api.maptiler.com/maps/satellite/{zoom}/{x}/{y}.jpg?key={key}"
WEATHERAPI_KEY = os.environ.get("WEATHERAPI_KEY", "").strip()

# ─────────────────────────────────────────────
# helpers
# ─────────────────────────────────────────────

def get_farm_safely(identifier):
    """البحث عن المزرعة عبر الـ ID أو رقم العقد"""
    try:
        doc_ref = DB.collection("farms").document(identifier).get()
        if doc_ref.exists:
            return doc_ref, "ID"
    except Exception as e:
        logger.error(f"Error fetching document: {e}")

    try:
        query = DB.collection("farms").where("contractNumber", "==", identifier).limit(1).get()
        docs = list(query)
        if docs:
            return docs[0], "contractNumber"
    except Exception as e:
        logger.error(f"Error querying contractNumber: {e}")

    return None, None


def _safe_float(v, default=0.0):
    try:
        if v is None:
            return default
        result = float(v)
        if math.isnan(result) or math.isinf(result):
            return default
        return result
    except Exception:
        return default


def _safe_int(v, default=0):
    try:
        if v is None:
            return default
        return int(v)
    except Exception:
        return default
    
INDEX_META = {
    "NDVI": ("الخضرة", "يقيس كثافة الغطاء النباتي وحيوية النمو بشكل عام."),
    "GNDVI": ("الخضرة الخضراء", "يركز أكثر على حساسية الكلوروفيل والنمو النشط."),
    "NDRE": ("حيوية الأوراق", "يفيد في كشف التغيرات المبكرة في نشاط الأوراق والتغذية."),
    "NDRE740": ("الحافة الحمراء 740", "مؤشر حساس للتغيرات الدقيقة في الكلوروفيل والنشاط الحيوي."),
    "MTCI": ("دليل الكلوروفيل", "يستخدم لتقدير الكلوروفيل وقد يساعد في قراءة الحالة التغذوية."),
    "NDMI": ("الرطوبة", "يعكس مستوى الرطوبة في المجموع الخضري واحتمال الإجهاد المائي."),
    "NDWI_Gao": ("مؤشر الماء", "يعكس محتوى الماء في النبات ويستخدم لدعم قراءة الإجهاد المائي."),
    "SIWSI1": ("إجهاد الماء 1", "يساعد في رصد الإجهاد المائي داخل الغطاء النباتي."),
    "SIWSI2": ("إجهاد الماء 2", "يعطي قراءة إضافية لحالة الماء والنشاط الحيوي."),
    "SRWI": ("نسبة الماء الطيفية", "مؤشر إضافي لدعم تقييم رطوبة النبات."),
    "NMDI": ("فرق الرطوبة الطبيعي", "يفيد في تقييم توازن الرطوبة والإجهاد المرتبط بها."),
}


def _enrich_indices_table(indices_table, ndvi=None, ndmi=None, ndre=None):
    """
    يضيف label و note للتقرير فقط بناءً على code.
    لا يغير التخزين في Firestore.
    """
    enriched = []

    for item in indices_table or []:
        code = str(item.get("code", "") or "").strip()
        if not code:
            continue

        label, note = INDEX_META.get(code, (code, "—"))

        enriched.append({
            "label": item.get("label") or label,
            "code": code,
            "value": item.get("value", 0),
            "note": item.get("note") or note,
        })

    if enriched:
        return enriched

    fallback_items = [
        ("NDVI", ndvi),
        ("NDMI", ndmi),
        ("NDRE", ndre),
    ]

    for code, metric in fallback_items:
        label, note = INDEX_META.get(code, (code, "—"))
        metric = metric or {}

        enriched.append({
            "label": label,
            "code": code,
            "value": f"{_safe_float(metric.get('val', 0), 0):.2f}",
            "note": note,
        })

    return enriched

def _load_local_image_as_data_uri(*relative_parts) -> str | None:
    try:
        path = os.path.join(os.path.dirname(__file__), *relative_parts)
        if not os.path.exists(path):
            return None
        with open(path, "rb") as img_file:
            img_b64 = base64.b64encode(img_file.read()).decode("utf-8")
        ext = os.path.splitext(path)[1].lower()
        mime = "image/png" if ext == ".png" else "image/jpeg" if ext in {".jpg", ".jpeg"} else "image/svg+xml" if ext == ".svg" else "application/octet-stream"
        return f"data:{mime};base64,{img_b64}"
    except Exception as e:
        logger.warning(f"Failed to load local image {'/'.join(relative_parts)}: {e}")
        return None

def _inline_svg_data_uri(svg_text: str) -> str:
    svg_b64 = base64.b64encode(svg_text.encode("utf-8")).decode("utf-8")
    return f"data:image/svg+xml;base64,{svg_b64}"


def _default_watermark_data_uri() -> str:
    svg = """
    <svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 220 220">
      <g fill="none" fill-rule="evenodd">
        <circle cx="110" cy="110" r="92" fill="#0f766e" opacity="0.10"/>
        <path d="M110 36 L152 78 L110 120 L68 78 Z" fill="#065f46" opacity="0.22"/>
        <path d="M110 58 C118 78, 118 102, 110 124 C102 102, 102 78, 110 58 Z" fill="#10b981" opacity="0.34"/>
      </g>
    </svg>
    """
    return _inline_svg_data_uri(svg)


def _color_for_pct(pct: float) -> str:
    if pct >= 75:
        return "#16a34a"
    if pct >= 45:
        return "#f59e0b"
    return "#dc2626"


def _gauge_svg(pct: float, color: str, size: int = 116) -> str:
    pct = max(0.0, min(100.0, float(pct)))
    r = 44
    cx = cy = size / 2
    circumference = 3.14159 * r
    stroke_dash = (pct / 100) * circumference
    stroke_gap = circumference - stroke_dash

    return f"""
<svg width="{size}" height="{size // 2 + 22}" viewBox="0 0 {size} {size // 2 + 22}">
  <path d="M {cx - r} {cy} A {r} {r} 0 0 1 {cx + r} {cy}"
        fill="none" stroke="#e5e7eb" stroke-width="10" stroke-linecap="round"/>
  <path d="M {cx - r} {cy} A {r} {r} 0 0 1 {cx + r} {cy}"
        fill="none" stroke="{color}" stroke-width="10" stroke-linecap="round"
        stroke-dasharray="{stroke_dash:.1f} {stroke_gap:.1f}"/>
  <text x="{cx}" y="{cy + 8}" text-anchor="middle"
        font-family="Cairo, sans-serif" font-size="20" font-weight="800" fill="{color}">
    {pct:.0f}%
  </text>
</svg>
""".strip()


def _trend_sparkline(values: list, color: str = "#2563eb", width: int = 235, height: int = 68) -> str:
    if not values:
        return ""

    vals = []
    for v in values:
        try:
            vals.append(float(v))
        except Exception:
            pass

    if len(vals) < 2:
        return ""

    mn, mx = min(vals), max(vals)
    rng = mx - mn if mx != mn else 1.0
    step = width / (len(vals) - 1)

    pts = []
    circles = []

    for i, v in enumerate(vals):
        x = i * step
        y = height - ((v - mn) / rng) * (height - 18) - 9
        pts.append(f"{x:.1f},{y:.1f}")
        circles.append(f'<circle cx="{x:.1f}" cy="{y:.1f}" r="3" fill="{color}" />')

    polyline = " ".join(pts)

    return f"""
<svg width="{width}" height="{height}" viewBox="0 0 {width} {height}">
  <line x1="0" y1="{height - 9}" x2="{width}" y2="{height - 9}" stroke="#dbeafe" stroke-width="1"/>
  <polyline points="{polyline}" fill="none" stroke="{color}" stroke-width="3"
            stroke-linejoin="round" stroke-linecap="round"/>
  {''.join(circles)}
</svg>
""".strip()


def _distribution_compare_svg(
    current_dist: dict,
    next_dist: dict,
    width: int = 300,
    height: int = 150
) -> str:
    """رسم أعمدة واضح بدون نصوص عربية داخل SVG لتجنب انعكاسها في PDF."""

    current_vals = [
        _safe_float(current_dist.get("Healthy_Pct", 0)),
        _safe_float(current_dist.get("Monitor_Pct", 0)),
        _safe_float(current_dist.get("Critical_Pct", 0)),
    ]
    next_vals = [
        _safe_float(next_dist.get("Healthy_Pct_next", 0)),
        _safe_float(next_dist.get("Monitor_Pct_next", 0)),
        _safe_float(next_dist.get("Critical_Pct_next", 0)),
    ]

    strong_colors = ["#22c55e", "#f59e0b", "#ef4444"]
    light_colors  = ["#bbf7d0", "#fde68a", "#fecaca"]

    pad_l, pad_r, pad_t, pad_b = 18, 18, 20, 26
    chart_h = height - pad_t - pad_b
    base_y = pad_t + chart_h
    bar_w = 22
    inner_gap = 7
    group_gap = 24
    x = pad_l
    parts = []

    for frac in [0.25, 0.5, 0.75, 1.0]:
        gy = pad_t + chart_h * (1 - frac)
        label = int(frac * 100)
        parts.append(f'<line x1="{pad_l}" y1="{gy:.1f}" x2="{width-pad_r}" y2="{gy:.1f}" stroke="#eef2f7" stroke-width="1"/>')
        parts.append(f'<text x="{pad_l-4}" y="{gy+3:.1f}" text-anchor="end" font-family="Arial, sans-serif" font-size="8" fill="#94a3b8">{label}</text>')

    parts.append(f'<line x1="{pad_l}" y1="{base_y}" x2="{width-pad_r}" y2="{base_y}" stroke="#cbd5e1" stroke-width="1.2"/>')

    for i in range(3):
        cur_v = max(0.0, min(100.0, current_vals[i]))
        nxt_v = max(0.0, min(100.0, next_vals[i]))
        cur_h = (cur_v / 100.0) * chart_h
        nxt_h = (nxt_v / 100.0) * chart_h
        cur_x = x
        nxt_x = x + bar_w + inner_gap
        cur_y = base_y - cur_h
        nxt_y = base_y - nxt_h

        parts.append(f'<rect x="{cur_x}" y="{cur_y:.1f}" width="{bar_w}" height="{cur_h:.1f}" rx="7" fill="{strong_colors[i]}"/>')
        parts.append(f'<rect x="{nxt_x}" y="{nxt_y:.1f}" width="{bar_w}" height="{nxt_h:.1f}" rx="7" fill="{light_colors[i]}" stroke="{strong_colors[i]}" stroke-width="0.9"/>')

        parts.append(f'<text x="{cur_x + bar_w/2:.1f}" y="{cur_y - 4:.1f}" text-anchor="middle" font-family="Arial, sans-serif" font-size="8" font-weight="700" fill="{strong_colors[i]}">{cur_v:.0f}%</text>')
        parts.append(f'<text x="{nxt_x + bar_w/2:.1f}" y="{nxt_y - 4:.1f}" text-anchor="middle" font-family="Arial, sans-serif" font-size="8" font-weight="700" fill="#64748b">{nxt_v:.0f}%</text>')

        x += (bar_w * 2 + inner_gap + group_gap)

    return f"""
<svg width="{width}" height="{height}" viewBox="0 0 {width} {height}" xmlns="http://www.w3.org/2000/svg">
  {''.join(parts)}
</svg>
""".strip()

def _mini_bar_svg(values: list[float], colors: list[str], width: int = 215, height: int = 84) -> str:
    """
    قراف أفقي صغير لعرض 3 قيم بسرعة بدون نصوص عربية داخل SVG.
    """
    if not values:
        return ""

    vals = [max(0.0, _safe_float(v, 0.0)) for v in values]
    mx = max(max(vals), 1.0)

    bar_h = 14
    gap = 10
    left = 8
    top = 8

    parts = []
    for i, v in enumerate(vals):
        y = top + i * (bar_h + gap)
        w = (v / mx) * (width - 40)
        parts.append(
            f'<rect x="{left}" y="{y}" width="{width-30}" height="{bar_h}" rx="7" fill="#edf2f7"/>'
        )
        parts.append(
            f'<rect x="{left}" y="{y}" width="{w:.1f}" height="{bar_h}" rx="7" fill="{colors[i]}"/>'
        )

    return f"""
<svg width="{width}" height="{height}" viewBox="0 0 {width} {height}">
  {''.join(parts)}
</svg>
""".strip()


def _multi_index_sparkline(
    multi_trend: dict,
    width: int = 430,
    height: int = 126,
) -> str:
    """رسم أوضح للمؤشرات الطيفية مع مساحة أكبر حتى لا يُقص داخل الكرت."""
    ndvi = [_safe_float(v) for v in (multi_trend.get("ndvi") or []) if v is not None]
    ndmi = [_safe_float(v) for v in (multi_trend.get("ndmi") or []) if v is not None]
    ndre = [_safe_float(v) for v in (multi_trend.get("ndre") or []) if v is not None]

    series = [
        (ndvi, "#16a34a", "NDVI"),
        (ndmi, "#2563eb", "NDMI"),
        (ndre, "#0f766e", "NDRE"),
    ]
    valid = [(vals, color, label) for vals, color, label in series if len(vals) >= 2]
    if not valid:
        return ""

    all_vals = [v for vals, _, _ in valid for v in vals]
    mn, mx = min(all_vals), max(all_vals)
    if mn == mx:
        mn -= 0.1
        mx += 0.1
    rng = mx - mn

    pad_l, pad_r, pad_t, pad_b = 34, 12, 10, 28
    chart_w = width - pad_l - pad_r
    chart_h = height - pad_t - pad_b

    parts = []

    for frac in [0.0, 0.25, 0.5, 0.75, 1.0]:
        y = pad_t + chart_h - (frac * chart_h)
        val = mn + frac * rng
        parts.append(f'<line x1="{pad_l}" y1="{y:.1f}" x2="{pad_l+chart_w}" y2="{y:.1f}" stroke="#e5e7eb" stroke-width="1"/>')
        parts.append(f'<text x="{pad_l-6}" y="{y+3:.1f}" text-anchor="end" font-family="Arial, sans-serif" font-size="8" fill="#94a3b8">{val:.2f}</text>')

    longest = max(len(vals) for vals, _, _ in valid)
    step = chart_w / max(longest - 1, 1)
    for i in range(longest):
        x = pad_l + i * step
        parts.append(f'<line x1="{x:.1f}" y1="{pad_t}" x2="{x:.1f}" y2="{pad_t+chart_h}" stroke="#f1f5f9" stroke-width="1"/>')

    for vals, color, label in valid:
        n = len(vals)
        local_step = chart_w / max(n - 1, 1)
        pts = []
        for i, v in enumerate(vals):
            x = pad_l + i * local_step
            y = pad_t + chart_h - ((v - mn) / rng) * chart_h
            pts.append((x, y))

        polyline = " ".join(f"{x:.1f},{y:.1f}" for x, y in pts)
        parts.append(f'<polyline points="{polyline}" fill="none" stroke="{color}" stroke-width="2.4" stroke-linejoin="round" stroke-linecap="round"/>')
        for x, y in pts:
            parts.append(f'<circle cx="{x:.1f}" cy="{y:.1f}" r="2.7" fill="{color}"/>')

    lx = pad_l
    for _, color, lbl in valid:
        parts.append(f'<rect x="{lx}" y="{height-14}" width="11" height="7" rx="3" fill="{color}"/>')
        parts.append(f'<text x="{lx+15}" y="{height-8}" font-family="Arial, sans-serif" font-size="8" fill="#475569">{lbl}</text>')
        lx += 62

    return f"""
<svg width="{width}" height="{height}" viewBox="0 0 {width} {height}" xmlns="http://www.w3.org/2000/svg">
  {''.join(parts)}
</svg>
""".strip()


def _flag_breakdown_rows(flag_counts: dict) -> list[dict]:
    """تفصيل الإشارات؛ هذا عدّ إشارات وليس عدّ بكسلات فريدة."""
    water = int(flag_counts.get("water_low", 0)) + int(flag_counts.get("water_below_025", 0))
    veg = int(flag_counts.get("ndvi_below_030", 0)) + int(flag_counts.get("ndvi_drop", 0))
    leaf = int(flag_counts.get("ndre_low", 0)) + int(flag_counts.get("ndre_below_035", 0))
    stress = int(flag_counts.get("siwsi_drop", 0)) + int(flag_counts.get("water_drop", 0))

    rows = [
        {"label": "إجهاد مائي", "value": water, "color": "#3b82f6"},
        {"label": "هبوط الخضرة", "value": veg, "color": "#22c55e"},
        {"label": "ضعف التغذية", "value": leaf, "color": "#0d9488"},
        {"label": "ضغط مائي مركب", "value": stress, "color": "#f59e0b"},
    ]

    max_value = max([r["value"] for r in rows] + [1])
    for row in rows:
        row["pct"] = round((row["value"] / max_value) * 100, 1)

    return rows

def _sector_distribution_rows(map_points: list) -> list[dict]:
    """تلخيص توزيع الحالات حسب القطاع المكاني داخل المزرعة."""
    points = []
    for p in map_points or []:
        try:
            lat = float(p.get("lat"))
            lng = float(p.get("lng"))
            s = int(p.get("currentStatus", 0))
            points.append((lat, lng, s))
        except Exception:
            continue

    if not points:
        return [
            {"label": "شمال", "healthy_pct": 0, "monitor_pct": 0, "critical_pct": 0, "total": 0},
            {"label": "جنوب", "healthy_pct": 0, "monitor_pct": 0, "critical_pct": 0, "total": 0},
            {"label": "شرق", "healthy_pct": 0, "monitor_pct": 0, "critical_pct": 0, "total": 0},
            {"label": "غرب", "healthy_pct": 0, "monitor_pct": 0, "critical_pct": 0, "total": 0},
        ]

    lats = [p[0] for p in points]
    lngs = [p[1] for p in points]
    lat_mid = (min(lats) + max(lats)) / 2
    lng_mid = (min(lngs) + max(lngs)) / 2

    buckets = {
        "شمال": {0: 0, 1: 0, 2: 0},
        "جنوب": {0: 0, 1: 0, 2: 0},
        "شرق": {0: 0, 1: 0, 2: 0},
        "غرب": {0: 0, 1: 0, 2: 0},
    }

    for lat, lng, s in points:
        if abs(lat - lat_mid) >= abs(lng - lng_mid):
            key = "شمال" if lat >= lat_mid else "جنوب"
        else:
            key = "شرق" if lng >= lng_mid else "غرب"
        buckets[key][2 if s == 2 else 1 if s == 1 else 0] += 1

    rows = []
    for label in ["شمال", "جنوب", "شرق", "غرب"]:
        counts = buckets[label]
        total = counts[0] + counts[1] + counts[2]
        if total == 0:
            rows.append({"label": label, "healthy_pct": 0, "monitor_pct": 0, "critical_pct": 0, "total": 0})
        else:
            rows.append({
                "label": label,
                "healthy_pct": round((counts[0] / total) * 100, 1),
                "monitor_pct": round((counts[1] / total) * 100, 1),
                "critical_pct": round((counts[2] / total) * 100, 1),
                "total": total,
            })
    return rows

def _normalize_polygon(poly: list | None) -> list[tuple[float, float]]:
    out = []
    if not poly:
        return out

    for item in poly:
        try:
            if isinstance(item, dict):
                lat = float(item.get("lat"))
                lng = float(item.get("lng"))
                out.append((lat, lng))
            elif isinstance(item, (list, tuple)) and len(item) >= 2:
                a = float(item[0])
                b = float(item[1])
                if abs(a) <= 90 and abs(b) <= 180:
                    out.append((a, b))
                else:
                    out.append((b, a))
        except Exception:
            continue

    return out

def _map_bounds(map_points: list, farm_polygon: list | None = None):
    norm_poly = _normalize_polygon(farm_polygon)

    lats, lngs = [], []

    for p in map_points or []:
        try:
            lats.append(float(p.get("lat")))
            lngs.append(float(p.get("lng")))
        except Exception:
            pass

    for lat, lng in norm_poly:
        lats.append(lat)
        lngs.append(lng)

    if not lats or not lngs:
        return None

    min_lat, max_lat = min(lats), max(lats)
    min_lng, max_lng = min(lngs), max(lngs)

    lat_pad = max((max_lat - min_lat) * 0.025, 0.00015)
    lng_pad = max((max_lng - min_lng) * 0.025, 0.00015)

    return {
        "min_lat": min_lat - lat_pad,
        "max_lat": max_lat + lat_pad,
        "min_lng": min_lng - lng_pad,
        "max_lng": max_lng + lng_pad,
    }




def _heatmap_svg(map_points: list, width: int = 505, height: int = 280, farm_polygon: list | None = None) -> dict:
    norm_poly = _normalize_polygon(farm_polygon)

    if not map_points and not norm_poly:
        return {"bg_data_uri": None, "overlay_svg": ""}

    bounds = _map_bounds(map_points, farm_polygon)
    if not bounds:
        return {"bg_data_uri": None, "overlay_svg": ""}

    stitched = _stitch_maptiler_tiles(bounds, width, height)
    bg_data_uri = stitched.get("bg_data_uri")
    meta = stitched.get("meta")

    color_map = {
        0: "#22c55e",
        1: "#f59e0b",
        2: "#ef4444",
    }

    def to_xy(lat, lng):
        if meta:
           gx, gy = _latlon_to_global_pixels(lat, lng, meta["zoom"])
           px = gx - meta["origin_x"]
           py = gy - meta["origin_y"]
           return round(px, 1), round(py, 1)

        min_lat = bounds["min_lat"]
        max_lat = bounds["max_lat"]
        min_lng = bounds["min_lng"]
        max_lng = bounds["max_lng"]
        lat_rng = max(max_lat - min_lat, 1e-6)
        lng_rng = max(max_lng - min_lng, 1e-6)
        pad = 8
        x = pad + ((lng - min_lng) / lng_rng) * (width - pad * 2)
        y = pad + ((max_lat - lat) / lat_rng) * (height - pad * 2)
        return round(x, 1), round(y, 1)

    poly_svg = ""
    if norm_poly:
        pts = []
        for lat, lng in norm_poly:
            x, y = to_xy(lat, lng)
            pts.append(f"{x},{y}")
        poly_svg = (
            f'<polygon points="{" ".join(pts)}" fill="white" fill-opacity="0.08" '
            f'stroke="#10b981" stroke-width="2.2" opacity="1"/>'
        )

    circles = []
    for pt in map_points or []:
        try:
            lat = float(pt.get("lat"))
            lng = float(pt.get("lng"))
            s = int(pt.get("currentStatus", 0))
            ps = int(pt.get("predictedStatus", s))
        except Exception:
            continue

        x, y = to_xy(lat, lng)

        if x < -20 or x > width + 20 or y < -20 or y > height + 20:
            continue

        r = 2.4 if s == 2 else 2.4 if s == 1 else 2.1
        fill = color_map.get(s, "#22c55e")

        ring = ""
        if ps != s:
            ring = (
                f'<circle cx="{x}" cy="{y}" r="{r + 1.4}" fill="none" '
                f'stroke="#2563eb" stroke-width="0.9" opacity="0.95"/>'
            )

        circles.append(
            ring +
            f'<circle cx="{x}" cy="{y}" r="{r}" fill="{fill}" stroke="white" stroke-width="0.5" opacity="0.98" />'
        )

    if bg_data_uri:
       bg_layer = f'<image href="{bg_data_uri}" x="0" y="0" width="{width}" height="{height}" preserveAspectRatio="none"/>'
    else:
       bg_layer = f'<rect x="0" y="0" width="{width}" height="{height}" fill="#e5efe8"/>'

    combined_svg = f"""<svg
    width="{width}"
    height="{height}"
    viewBox="0 0 {width} {height}"
    xmlns="http://www.w3.org/2000/svg">
  {bg_layer}
  {poly_svg}
  {''.join(circles)}
</svg>""".strip()

    logger.info(
    "Heatmap debug | points=%s | bg=%s | meta=%s",
    len(map_points or []),
    bool(bg_data_uri),
    meta,
)

    return {"bg_data_uri": None, "overlay_svg": combined_svg}


def _footer_logo_html(logo_data_uri: str | None) -> str:
    if logo_data_uri:
        return (
            f'<img src="{logo_data_uri}" alt="Saaf" '
            f'style="height:42px; width:auto; max-width:150px; display:block; object-fit:contain;" />'
        )
    return '<div style="font-size:16px;font-weight:900;color:#6ee7b7;letter-spacing:1px;">سعف</div>'

def _delta_badge_html(delta_pct: float) -> str:
    delta_pct = _safe_float(delta_pct, 0.0)

    if abs(delta_pct) < 0.5:
        return '<span style="color:#64748b;">بدون تغير ملحوظ</span>'

    if delta_pct > 0:
        return f'<span style="color:#16a34a;">▲ {abs(delta_pct):.1f}%</span>'

    return f'<span style="color:#dc2626;">▼ {abs(delta_pct):.1f}%</span>'


def _metric_verdict(code: str, value: float):
    v = _safe_float(value, 0.0)

    if code == "NDVI":
        if v >= 0.60:
            return "جيدة", "#dcfce7", "#166534"
        if v >= 0.35:
            return "متوسطة", "#fef3c7", "#92400e"
        return "منخفضة", "#fee2e2", "#b91c1c"

    if code == "NDMI":
        if v >= 0.30:
            return "جيدة", "#dbeafe", "#1d4ed8"
        if v >= 0.10:
            return "متوسطة", "#fef3c7", "#92400e"
        return "منخفضة", "#fee2e2", "#b91c1c"

    if code == "NDRE":
        if v >= 0.45:
            return "جيدة", "#dcfce7", "#166534"
        if v >= 0.25:
            return "متوسطة", "#fef3c7", "#92400e"
        return "منخفضة", "#fee2e2", "#b91c1c"

    return "—", "#f1f5f9", "#475569"

def style_chart_axes(chart, *, y_title=None, x_title=None, show_values=False, is_percent=False):
    try:
        if y_title is not None:
            chart.y_axis.title = y_title
        if x_title is not None:
            chart.x_axis.title = x_title

        chart.y_axis.delete = False
        chart.x_axis.delete = False

        chart.y_axis.tickLblPos = "nextTo"
        chart.x_axis.tickLblPos = "low"

        if is_percent:
            chart.y_axis.scaling.min = 0
            chart.y_axis.scaling.max = 100
            chart.y_axis.majorUnit = 20

        if show_values:
            chart.dLbls = DataLabelList()
            chart.dLbls.showVal = True
            chart.dLbls.showLegendKey = False
            chart.dLbls.showCatName = False
            chart.dLbls.showSerName = False
            chart.dLbls.showPercent = False
    except Exception:
        pass

# ─────────────────────────────────────────────
# PDF Generation — weasyprint
# ─────────────────────────────────────────────
def generate_pdf_report(export_data: dict, farm_id: str, farm_doc: dict | None = None) -> str:
    try:
        from weasyprint import HTML, CSS
    except ImportError:
        raise RuntimeError("weasyprint غير مثبتة. أضف 'weasyprint' لـ requirements.txt")

    farm_doc = farm_doc or {}

    header = export_data.get("header", {})
    dist = export_data.get("distribution", {})
    biometrics = export_data.get("biometrics", {})
    forecast = export_data.get("forecast", {})
    forecast_next = export_data.get("forecast_next_week", {})
    top_action = export_data.get("top_action") or {}
    health_root = farm_doc.get("health", {}) if isinstance(farm_doc.get("health"), dict) else {}
    map_points = (
    export_data.get("health_map_points")
    or export_data.get("health_map")
    or farm_doc.get("healthMap", [])
    or health_root.get("health_map", [])
    or []
     )    
    farm_poly = export_data.get("farm_polygon", []) or farm_doc.get("polygon", [])
    risk_drivers = (export_data.get("risk_drivers", []) or [])[:3]
    hotspots = (export_data.get("hotspots_table", []) or [])[:3]
    indices_table = export_data.get("indices_table", []) or []
    climate = export_data.get("climate", {}) or {}
    alert_context = export_data.get("alert_context", {}) or {}
    multi_trend = export_data.get("multi_trend", {}) or {}
    owner_name = (
        export_data.get("owner_name")
        or header.get("owner_name")
        or farm_doc.get("ownerName")
        or "—"
    )

    total_palms = (
        header.get("total_palms")
        or export_data.get("total_palms")
        or export_data.get("finalCount")
        or farm_doc.get("finalCount")
        or farm_doc.get("palmCount")
        or farm_doc.get("totalPalms")
        or 0
    )
    total_palms = _safe_int(total_palms, 0)

    wellness = float(export_data.get("wellness_score", dist.get("Healthy_Pct", 0)))

    healthy_pct = _safe_float(dist.get("Healthy_Pct", 0))
    monitor_pct = _safe_float(dist.get("Monitor_Pct", 0))
    critical_pct = _safe_float(dist.get("Critical_Pct", 0))

    gauge_color = _color_for_pct(wellness)
    gauge_svg = _gauge_svg(wellness, gauge_color, size=124)
    sparkline = _trend_sparkline(forecast.get("trend_data", []), color="#2563eb")
    map_result = _heatmap_svg(map_points, width=490, height=176, farm_polygon=farm_poly)
    map_bg_uri = map_result["bg_data_uri"]
    map_svg    = map_result["overlay_svg"]
    compare_svg = _distribution_compare_svg(dist, forecast_next)
    drivers_bar_svg = _mini_bar_svg(
        [r.get("count", 0) for r in risk_drivers[:3]],
        ["#0ea5e9", "#f59e0b", "#ef4444"]
    )

    multi_sparkline_svg = _multi_index_sparkline(multi_trend)
    flag_rows = _flag_breakdown_rows(alert_context.get("flag_counts", {}))
    total_flag_signals = sum(int(r.get("value", 0) or 0) for r in flag_rows)
    sector_rows = _sector_distribution_rows(map_points)
    total_map_points = len([p for p in (map_points or []) if p.get("lat") is not None and p.get("lng") is not None])

    rain_mm       = _safe_float(climate.get("rain_mm", 0))
    t_mean        = _safe_float(climate.get("t_mean", 0))
    rpw_score     = _safe_float(climate.get("rpw_score", 0))

    total_pixels = int(alert_context.get("total_pixels", 0) or climate.get("total_pixels", 0) or 0)
    pixels_with_any_flag = int(alert_context.get("pixels_with_any_flag", 0) or 0)
    signal_note = alert_context.get("signal_note") or "قد تُسجَّل أكثر من إشارة للبكسل الواحد."

    flag_ratio = (pixels_with_any_flag / total_pixels) if total_pixels else 0.0

    if critical_pct >= 10:
       rpw_diagnosis = "تحتاج الحالة متابعة عاجلة"
       rpw_diagnosis_sub = "توجد نسبة حرجة ملحوظة في التوزيع العام للمزرعة."
    elif critical_pct >= 5 or monitor_pct >= 25 or flag_ratio >= 0.20:
       rpw_diagnosis = "تحتاج الحالة متابعة"
       rpw_diagnosis_sub = "هناك مناطق متأثرة تستلزم مراقبة أقرب خلال الفترة القادمة."
    else:
       rpw_diagnosis = "الحالة مستقرة حاليًا"
       rpw_diagnosis_sub = "الوضع العام جيد مع ملاحظات محدودة لا تستدعي تصعيدًا."

    if rpw_score >= 0.66:
        rpw_label = "مرتفع"
        rpw_color = "#dc2626"
    elif rpw_score >= 0.40:
        rpw_label = "متوسط"
        rpw_color = "#f59e0b"
    else:
        rpw_label = "منخفض"
        rpw_color = "#16a34a"
    
    ndvi = biometrics.get("ndvi", {})
    ndmi = biometrics.get("ndmi", {})
    ndre = biometrics.get("ndre", {})

    ndvi_state, ndvi_bg, ndvi_fg = _metric_verdict("NDVI", ndvi.get("val", 0))
    ndmi_state, ndmi_bg, ndmi_fg = _metric_verdict("NDMI", ndmi.get("val", 0))
    ndre_state, ndre_bg, ndre_fg = _metric_verdict("NDRE", ndre.get("val", 0))

    if healthy_pct >= 85 and critical_pct < 5:
        wellness_text = "الحالة العامة جيدة جدًا"
    elif healthy_pct >= 65 and critical_pct < 10:
        wellness_text = "الحالة العامة جيدة"
    elif critical_pct >= 10:
        wellness_text = "الحالة العامة تتطلب تدخلًا أسرع"
    else:
        wellness_text = "الحالة العامة تحتاج متابعة"

    wellness_desc = (
        "يعرض هذا التقرير الوضع الحالي للمزرعة استنادًا إلى المؤشرات الطيفية "
        "وتوزيع النقاط المتأثرة داخل حدود المزرعة."
    )

    report_date = header.get("date") or datetime.now().strftime("%Y-%m-%d")
    farm_name = header.get("name") or export_data.get("farmName") or farm_doc.get("farmName") or "—"
    farm_area = header.get("area") or export_data.get("farmSize") or farm_doc.get("farmSize") or "—"
    contract_number = header.get("contract_number") or farm_doc.get("contractNumber") or "—"
    region = header.get("city") or farm_doc.get("region") or "—"

    executive_status = export_data.get("executive_status", "ملخص الحالة")
    executive_summary = export_data.get("executive_summary", "—")
    executive_next_step = export_data.get("executive_next_step", top_action.get("title_ar", "—"))

    forecast_text = forecast.get("text", "—")
    forecast_summary = {
        "healthy": f"{_safe_float(forecast_next.get('Healthy_Pct_next', healthy_pct), 0):.1f}%",
        "monitor": f"{_safe_float(forecast_next.get('Monitor_Pct_next', monitor_pct), 0):.1f}%",
        "critical": f"{_safe_float(forecast_next.get('Critical_Pct_next', critical_pct), 0):.1f}%",
    }
    forecast_change_text = _delta_badge_html(_safe_float(forecast_next.get("ndvi_delta_next_mean"), 0) * 100.0)

    indices_table = _enrich_indices_table(
        indices_table,
        ndvi=ndvi,
        ndmi=ndmi,
        ndre=ndre,
    )


    logo_data_uri = _load_local_image_as_data_uri("static", "images", "saaf_logo.png")
    palm_icon_data_uri = _load_local_image_as_data_uri("static", "images", "PalmIcon.png")

    if not logo_data_uri:
        logger.warning("saaf_logo.png was not loaded, using inline fallback watermark.")
    if not palm_icon_data_uri:
        logger.warning("PalmIcon.png was not loaded, using fallback watermark.")

    watermark_data_uri = palm_icon_data_uri or logo_data_uri or _default_watermark_data_uri()
    footer_logo_html = _footer_logo_html(logo_data_uri)

    html_content = render_template(
        "reports/farm_report.html",
        farm_id=farm_id,
        farm_name=farm_name,
        farm_area=farm_area,
        total_palms=total_palms,
        report_date=report_date,
        contract_number=contract_number,
        region=region,
        owner_name=owner_name,

        logo_data_uri=logo_data_uri,
        palm_icon_data_uri=watermark_data_uri,
        footer_logo_html=footer_logo_html,

        executive_status=executive_status,
        executive_summary=executive_summary,
        executive_next_step=executive_next_step,

        gauge_svg=gauge_svg,
        gauge_color=gauge_color,
        wellness_text=wellness_text,
        wellness_desc=wellness_desc,
        healthy_pct=healthy_pct,
        monitor_pct=monitor_pct,
        critical_pct=critical_pct,

        map_svg=map_svg,
        map_bg_uri=map_bg_uri,
        compare_labels=["سليم", "متابعة", "حرج"],
        total_flag_signals=total_flag_signals,
        sector_rows=sector_rows,
        total_map_points=total_map_points,

        compare_svg=compare_svg,
        drivers_bar_svg=drivers_bar_svg,
        multi_sparkline_svg=multi_sparkline_svg,
        flag_rows=flag_rows,
        stress_pixels=f"{pixels_with_any_flag:,}".replace(",", "،"),
        stress_pixels_note=signal_note,
        trend_start_label=(multi_trend.get("dates") or ["الأقدم"])[0] if (multi_trend.get("dates") or []) else "الأقدم",
        trend_end_label=(multi_trend.get("dates") or ["الأحدث"])[-1] if (multi_trend.get("dates") or []) else "الأحدث",

        rain_mm=f"{rain_mm:.1f}",
        t_mean=f"{t_mean:.1f}",
        rpw_score=f"{rpw_score:.2f}",
        rpw_label=rpw_label,
        rpw_color=rpw_color,
        total_pixels=f"{total_pixels:,}".replace(",", "،"),

        ndvi_val=f"{_safe_float(ndvi.get('val', 0), 0):.2f}",
        ndmi_val=f"{_safe_float(ndmi.get('val', 0), 0):.2f}",
        ndre_val=f"{_safe_float(ndre.get('val', 0), 0):.2f}",

        ndvi_state=ndvi_state,
        ndvi_bg=ndvi_bg,
        ndvi_fg=ndvi_fg,
        ndmi_state=ndmi_state,
        ndmi_bg=ndmi_bg,
        ndmi_fg=ndmi_fg,
        ndre_state=ndre_state,
        ndre_bg=ndre_bg,
        ndre_fg=ndre_fg,

        ndvi_delta_badge=_delta_badge_html(_safe_float(ndvi.get("delta", 0), 0)),
        ndmi_delta_badge=_delta_badge_html(_safe_float(ndmi.get("delta", 0), 0)),
        ndre_delta_badge=_delta_badge_html(_safe_float(ndre.get("delta", 0), 0)),

        indices_table=indices_table,
        forecast_text=forecast_text,
        forecast_change_text=forecast_change_text,
        sparkline=sparkline,
        forecast_summary=forecast_summary,

        risk_drivers=risk_drivers,
        hotspots=hotspots,
        rpw_diagnosis=rpw_diagnosis,
        rpw_diagnosis_sub=rpw_diagnosis_sub,
    )

    output_path = f"/tmp/saaf_report_{farm_id}.pdf"

    HTML(string=html_content).write_pdf(
        output_path,
        stylesheets=[CSS(string="@page { size: A4; margin: 0; }")]
    )

    return output_path


# ─────────────────────────────────────────────
# Excel Generation — openpyxl
# ─────────────────────────────────────────────
def generate_excel_report(export_data: dict, farm_id: str) -> str:
    import os
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, PieChart, LineChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.chart.label import DataLabelList
    from openpyxl.chart.series import DataPoint

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "الملخص التنفيذي"
    ws.sheet_view.rightToLeft = True

    # ─────────────────────────────────────────────
    # helpers
    # ─────────────────────────────────────────────
    GREEN_DARK = "064E3B"
    GREEN_MID = "10B981"
    GREEN_LIGHT = "D1FAE5"
    GREEN_SOFT = "ECFDF5"

    ORANGE = "F59E0B"
    ORANGE_LIGHT = "FEF3C7"
    ORANGE_SOFT = "FFF7ED"

    RED = "EF4444"
    RED_LIGHT = "FEE2E2"
    RED_SOFT = "FEF2F2"

    BLUE = "2563EB"
    BLUE_LIGHT = "DBEAFE"
    BLUE_SOFT = "EFF6FF"

    SLATE = "334155"
    SLATE_2 = "64748B"
    SLATE_LIGHT = "F8FAFC"
    WHITE = "FFFFFF"
    BORDER = "E2E8F0"
    GOLD = "B45309"

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def side(color=BORDER, style="thin"):
        return Side(style=style, color=color)

    def border(color=BORDER, style="thin"):
        s = side(color, style)
        return Border(left=s, right=s, top=s, bottom=s)

    def font(size=11, bold=False, color="1E293B"):
        return Font(name="Cairo", size=size, bold=bold, color=color)

    def align_center(wrap=False):
        return Alignment(horizontal="center", vertical="center", wrap_text=wrap, readingOrder=2)

    def align_right(wrap=False):
        return Alignment(horizontal="right", vertical="center", wrap_text=wrap, readingOrder=2)

    def set_cell(cell, value=None, *, f=None, bg=None, al=None, bd=None):
        if value is not None:
            cell.value = value
        if f is not None:
            cell.font = f
        if bg is not None:
            cell.fill = fill(bg)
        if al is not None:
            cell.alignment = al
        if bd is not None:
            cell.border = bd

    def merge_value(ws_, cell_range, value, *, f=None, bg=None, al=None, bd=None):
        ws_.merge_cells(cell_range)
        cell = ws_[cell_range.split(":")[0]]
        set_cell(cell, value, f=f, bg=bg, al=al, bd=bd)
        return cell

    def pct(v):
        return f"{_safe_float(v, 0.0):.1f}%"

    def num(v, digits=2):
        return f"{_safe_float(v, 0.0):.{digits}f}"

    def delta_abs(v, digits=3):
        return f"{_safe_float(v, 0.0):+.{digits}f}"

    def safe_dict(v):
        return v if isinstance(v, dict) else {}

    def safe_list(v):
        return v if isinstance(v, list) else []

    def status_fill(status_text: str):
        s = str(status_text or "").strip()
        if "حرج" in s:
            return RED_SOFT
        if "متابعة" in s:
            return ORANGE_SOFT
        if "سليم" in s:
            return GREEN_SOFT
        return None

    def priority_fill(priority_text: str):
        p = str(priority_text or "").strip()
        if "مرتفعة" in p or "عاج" in p or "حرج" in p:
            return RED_LIGHT
        if "متوسطة" in p:
            return ORANGE_LIGHT
        if "منخفضة" in p:
            return GREEN_SOFT
        return SLATE_LIGHT

    def metric_fill(metric_code, value):
        v = _safe_float(value, 0.0)
        if metric_code == "NDVI":
            if v >= 0.60:
                return GREEN_SOFT, "166534"
            if v >= 0.35:
                return ORANGE_LIGHT, "92400E"
            return RED_LIGHT, "991B1B"
        if metric_code == "NDMI":
            if v >= 0.30:
                return BLUE_SOFT, "1D4ED8"
            if v >= 0.10:
                return ORANGE_LIGHT, "92400E"
            return RED_LIGHT, "991B1B"
        if metric_code == "NDRE":
            if v >= 0.45:
                return GREEN_SOFT, "166534"
            if v >= 0.25:
                return ORANGE_LIGHT, "92400E"
            return RED_LIGHT, "991B1B"
        return SLATE_LIGHT, SLATE

    def decode_status(code):
        try:
            code = int(code)
        except Exception:
            return "—"
        if code == 0:
            return "سليم"
        if code == 1:
            return "متابعة"
        if code == 2:
            return "حرج"
        return "—"

    def add_logo(target_ws):
        logo_path = os.path.join(os.path.dirname(__file__), "static", "images", "saaf_logo.png")
        if os.path.exists(logo_path):
            try:
                img = XLImage(logo_path)
                img.width = 115
                img.height = 42
                target_ws.add_image(img, "A1")
            except Exception:
                pass

    def draw_kpi_box(target_ws, start_col, row_top, title, value, subtitle="", tone="green"):
        tones = {
            "green":  (GREEN_SOFT, GREEN_DARK, "166534"),
            "orange": (ORANGE_LIGHT, ORANGE, "92400E"),
            "red":    (RED_LIGHT, RED, "991B1B"),
            "blue":   (BLUE_LIGHT, BLUE, "1D4ED8"),
            "slate":  (SLATE_LIGHT, SLATE, "334155"),
            "gold":   ("FFF7ED", GOLD, "92400E"),
        }
        bg, bd, fg = tones.get(tone, tones["slate"])
        end_col = start_col + 1

        for r in range(row_top, row_top + 3):
            for c in range(start_col, end_col + 1):
                cell = target_ws.cell(r, c)
                cell.fill = fill(bg)
                cell.border = border(bd)

        target_ws.merge_cells(start_row=row_top, start_column=start_col, end_row=row_top, end_column=end_col)
        target_ws.merge_cells(start_row=row_top + 1, start_column=start_col, end_row=row_top + 1, end_column=end_col)
        target_ws.merge_cells(start_row=row_top + 2, start_column=start_col, end_row=row_top + 2, end_column=end_col)

        set_cell(
            target_ws.cell(row_top, start_col),
            title,
            f=font(10, True, SLATE_2),
            al=align_center(True)
        )
        set_cell(
            target_ws.cell(row_top + 1, start_col),
            value,
            f=font(16, True, fg),
            al=align_center(True)
        )
        set_cell(
            target_ws.cell(row_top + 2, start_col),
            subtitle,
            f=font(9, False, SLATE_2),
            al=align_center(True)
        )

    def write_label_value_table(target_ws, start_row, title, rows):
        merge_value(
            target_ws, f"A{start_row}:H{start_row}", title,
            f=font(12, True, WHITE), bg=GREEN_MID, al=align_center(), bd=border()
        )
        r = start_row + 1
        for label, value in rows:
            merge_value(
                target_ws, f"A{r}:B{r}", label,
                f=font(10, True, GREEN_DARK), bg=GREEN_LIGHT, al=align_center(True), bd=border()
            )
            merge_value(
                target_ws, f"C{r}:H{r}", value,
                f=font(10, False, "1E293B"), bg=WHITE, al=align_right(True), bd=border()
            )
            r += 1
        return r

    def write_table(target_ws, start_row, title, headers, rows, *, title_color=GREEN_MID):
        last_col = len(headers)
        merge_value(
            target_ws,
            f"A{start_row}:{get_column_letter(last_col)}{start_row}",
            title,
            f=font(12, True, WHITE),
            bg=title_color,
            al=align_center(),
            bd=border()
        )

        hdr = start_row + 1
        for i, h in enumerate(headers, start=1):
            set_cell(
                target_ws.cell(hdr, i),
                h,
                f=font(10, True, SLATE),
                bg=SLATE_LIGHT,
                al=align_center(True),
                bd=border()
            )

        r = hdr + 1
        for row_data in rows:
            row_bg = None
            if len(row_data) > 0:
                row_bg = status_fill(str(row_data[0])) if headers and headers[0] == "الحالة" else None

            for i, val in enumerate(row_data, start=1):
                cell = target_ws.cell(r, i)
                set_cell(
                    cell,
                    val,
                    f=font(10, False, "1E293B"),
                    bg=row_bg,
                    al=align_center(True) if i != len(row_data) else align_right(True),
                    bd=border()
                )
            r += 1
        return r

    header = safe_dict(export_data.get("header"))
    dist = safe_dict(export_data.get("distribution"))
    forecast = safe_dict(export_data.get("forecast"))
    forecast_next = safe_dict(export_data.get("forecast_next_week"))
    biometrics = safe_dict(export_data.get("biometrics"))
    climate = safe_dict(export_data.get("climate"))
    alert_context = safe_dict(export_data.get("alert_context"))
    risk_drivers = safe_list(export_data.get("risk_drivers"))
    hotspots = safe_list(export_data.get("hotspots_table"))
    indices_table = safe_list(export_data.get("indices_table"))
    multi_trend = safe_dict(export_data.get("multi_trend"))
    map_points = safe_list(export_data.get("health_map_points"))

    farm_name = header.get("name") or export_data.get("farmName") or "—"
    contract_number = header.get("contract_number") or "—"
    region = header.get("city") or "—"
    owner_name = export_data.get("owner_name") or header.get("owner_name") or "—"
    farm_area = str(header.get("area") or export_data.get("farmSize") or "—")
    total_palms = _safe_int(
        header.get("total_palms")
        or export_data.get("palm_count")
        or export_data.get("finalCount")
        or 0
    )
    report_date = header.get("date") or datetime.now().strftime("%Y-%m-%d")

    healthy_pct = _safe_float(dist.get("Healthy_Pct", 0))
    monitor_pct = _safe_float(dist.get("Monitor_Pct", 0))
    critical_pct = _safe_float(dist.get("Critical_Pct", 0))
    wellness = _safe_float(export_data.get("wellness_score", healthy_pct), 0)

    executive_status = export_data.get("executive_status") or "ملخص الحالة"
    executive_summary = export_data.get("executive_summary") or "—"
    executive_next_step = export_data.get("executive_next_step") or "—"

    rain_mm = _safe_float(climate.get("rain_mm", 0))
    t_mean = _safe_float(climate.get("t_mean", 0))
    total_pixels = _safe_int(climate.get("total_pixels", 0))
    rpw_score = _safe_float(climate.get("rpw_score", 0))

    pixels_with_any_flag = _safe_int(alert_context.get("pixels_with_any_flag", 0))
    signal_note = alert_context.get("signal_note") or "—"
    flag_counts = safe_dict(alert_context.get("flag_counts"))
    rule_counts = safe_dict(alert_context.get("rule_counts"))

    ndvi = safe_dict(biometrics.get("ndvi"))
    ndmi = safe_dict(biometrics.get("ndmi"))
    ndre = safe_dict(biometrics.get("ndre"))

    ndvi_val = _safe_float(ndvi.get("val", 0))
    ndmi_val = _safe_float(ndmi.get("val", 0))
    ndre_val = _safe_float(ndre.get("val", 0))

    ndvi_delta = _safe_float(ndvi.get("delta", 0))
    ndmi_delta = _safe_float(ndmi.get("delta", 0))
    ndre_delta = _safe_float(ndre.get("delta", 0))

    forecast_text = forecast.get("text") or "—"
    healthy_next = _safe_float(forecast_next.get("Healthy_Pct_next", healthy_pct))
    monitor_next = _safe_float(forecast_next.get("Monitor_Pct_next", monitor_pct))
    critical_next = _safe_float(forecast_next.get("Critical_Pct_next", critical_pct))
    ndvi_delta_next = _safe_float(forecast_next.get("ndvi_delta_next_mean", 0))
    ndmi_delta_next = _safe_float(forecast_next.get("ndmi_delta_next_mean", 0))

    # ─────────────────────────────────────────────
    # Sheet 1: الملخص التنفيذي
    # ─────────────────────────────────────────────
    add_logo(ws)
    for col, width in {
        "A": 16, "B": 16, "C": 16, "D": 16,
        "E": 16, "F": 16, "G": 16, "H": 16
    }.items():
        ws.column_dimensions[col].width = width

    merge_value(
        ws, "A1:H1", "سعف — تقرير تحليل صحة المزرعة",
        f=font(16, True, WHITE), bg=GREEN_DARK, al=align_center(), bd=border(GREEN_DARK)
    )
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 10

    merge_value(
        ws, "A3:H3", "لوحة الملخص التنفيذي",
        f=font(12, True, WHITE), bg=GREEN_MID, al=align_center(), bd=border()
    )

    draw_kpi_box(ws, 1, 4, "مؤشر الحالة", pct(wellness), "الوضع العام", "green")
    draw_kpi_box(ws, 3, 4, "السليم", pct(healthy_pct), "الحالة الحالية", "green")
    draw_kpi_box(ws, 5, 4, "المتابعة", pct(monitor_pct), "تحتاج مراقبة", "orange")
    draw_kpi_box(ws, 7, 4, "الحرج", pct(critical_pct), "تحتاج تدخل", "red")

    draw_kpi_box(ws, 1, 8, "الأمطار", num(rain_mm, 1), "ملم", "blue")
    draw_kpi_box(ws, 3, 8, "متوسط الحرارة", num(t_mean, 1), "°م", "gold")
    draw_kpi_box(ws, 5, 8, "البكسلات المحللة", str(total_pixels), "آخر قراءة", "slate")
    draw_kpi_box(ws, 7, 8, "البكسلات المتأثرة", str(pixels_with_any_flag), "Monitor/Critical", "orange")

    merge_value(
        ws, "A12:H12", executive_status,
        f=font(13, True, GREEN_DARK), bg="F0FDF4", al=align_center(True), bd=border(GREEN_MID)
    )
    merge_value(
        ws, "A13:H15", executive_summary,
        f=font(11, False, "1E293B"), bg=WHITE, al=align_right(True), bd=border()
    )
    merge_value(
        ws, "A16:H16", f"الإجراء المقترح: {executive_next_step}",
        f=font(11, True, GREEN_DARK), bg=GREEN_LIGHT, al=align_center(True), bd=border()
    )

    info_rows = [
        ("اسم المزرعة", farm_name),
        ("رقم العقد", contract_number),
        ("المنطقة", region),
        ("اسم المالك", owner_name),
        ("المساحة", farm_area),
        ("عدد النخيل", str(total_palms)),
        ("تاريخ التقرير", report_date),
        ("RPW Score", num(rpw_score, 2)),
    ]
    row = write_label_value_table(ws, 18, "بيانات المزرعة", info_rows)

    forecast_rows = [
        ("نص التوقع", forecast_text),
        ("السليم المتوقع", pct(healthy_next)),
        ("المتابعة المتوقعة", pct(monitor_next)),
        ("الحرج المتوقع", pct(critical_next)),
        ("Delta NDVI القادم", delta_abs(ndvi_delta_next, 3)),
        ("Delta NDMI القادم", delta_abs(ndmi_delta_next, 3)),
        ("ملاحظة", signal_note),
    ]
    row = write_label_value_table(ws, row + 1, "توقع الأسبوع القادم", forecast_rows)

    data_row = row + 3
    ws.cell(data_row, 1, "الفئة")
    ws.cell(data_row, 2, "الحالي")
    ws.cell(data_row, 3, "المتوقع")
    ws.cell(data_row + 1, 1, "سليم")
    ws.cell(data_row + 1, 2, healthy_pct)
    ws.cell(data_row + 1, 3, healthy_next)
    ws.cell(data_row + 2, 1, "متابعة")
    ws.cell(data_row + 2, 2, monitor_pct)
    ws.cell(data_row + 2, 3, monitor_next)
    ws.cell(data_row + 3, 1, "حرج")
    ws.cell(data_row + 3, 2, critical_pct)
    ws.cell(data_row + 3, 3, critical_next)

    for r in range(data_row, data_row + 5):
        ws.row_dimensions[r].hidden = True

    bar = BarChart()
    bar.type = "col"
    bar.style = 10
    bar.title = "الحالي مقابل المتوقع"
    bar.height = 8
    bar.width = 12
    bar.add_data(
        Reference(ws, min_col=2, max_col=3, min_row=data_row, max_row=data_row + 3),
        titles_from_data=True
    )
    bar.set_categories(
        Reference(ws, min_col=1, min_row=data_row + 1, max_row=data_row + 3)
    )
    style_chart_axes(
        bar,
        y_title="النسبة %",
        x_title="الحالة",
        show_values=True,
        is_percent=True,
    )
    ws.add_chart(bar, "E37")

    # ─────────────────────────────────────────────
    # Sheet 2: المؤشرات والتحليل
    # ─────────────────────────────────────────────
    ws2 = wb.create_sheet("المؤشرات والتحليل")
    ws2.sheet_view.rightToLeft = True
    add_logo(ws2)
    for col, width in {
        "A": 22, "B": 14, "C": 14, "D": 46, "E": 14, "F": 14, "G": 14, "H": 14
    }.items():
        ws2.column_dimensions[col].width = width

    merge_value(
        ws2, "A1:H1", f"المؤشرات والتحليل — {farm_name}",
        f=font(16, True, WHITE), bg=GREEN_DARK, al=align_center(), bd=border(GREEN_DARK)
    )

    merge_value(
        ws2, "A3:H3", "ملخص المؤشرات الرئيسية",
        f=font(12, True, WHITE), bg=GREEN_MID, al=align_center(), bd=border()
    )

    draw_kpi_box(ws2, 1, 4, "NDVI", num(ndvi_val, 2), delta_abs(ndvi_delta, 3), "green")
    draw_kpi_box(ws2, 3, 4, "NDMI", num(ndmi_val, 2), delta_abs(ndmi_delta, 3), "blue")
    draw_kpi_box(ws2, 5, 4, "NDRE", num(ndre_val, 2), delta_abs(ndre_delta, 3), "gold")
    draw_kpi_box(
        ws2, 7, 4, "RPW Score", num(rpw_score, 2),
        "قراءة إجهاد", "red" if rpw_score >= 0.66 else "orange" if rpw_score >= 0.40 else "green"
    )

    indices_table = _enrich_indices_table(
        indices_table,
        ndvi=ndvi,
        ndmi=ndmi,
        ndre=ndre,
    )

    index_rows = []
    for item in indices_table:
        code = item.get("code", "—")
        value = _safe_float(item.get("value", 0))
        index_rows.append([
            item.get("label", "—"),
            code,
            value,
            item.get("note", "—"),
        ])

    if not index_rows:
        index_rows = [["—", "—", "—", "لا توجد بيانات مؤشرات"]]

    row2 = 9
    end2 = write_table(
        ws2, row2, "جدول المؤشرات الطيفية",
        ["المؤشر", "الرمز", "القيمة", "التفسير"],
        index_rows,
        title_color=GREEN_MID
    )

    for r in range(row2 + 2, end2):
        code = str(ws2.cell(r, 2).value)
        val = ws2.cell(r, 3).value
        bg, fg = metric_fill(code, val)
        ws2.cell(r, 3).fill = fill(bg)
        ws2.cell(r, 3).font = font(10, True, fg)

    delta_rows = [
        ["NDVI", num(ndvi_val, 2), delta_abs(ndvi_delta, 3)],
        ["NDMI", num(ndmi_val, 2), delta_abs(ndmi_delta, 3)],
        ["NDRE", num(ndre_val, 2), delta_abs(ndre_delta, 3)],
        ["NDVI القادم", "—", delta_abs(ndvi_delta_next, 3)],
        ["NDMI القادم", "—", delta_abs(ndmi_delta_next, 3)],
    ]
    row2 = end2 + 1
    end_delta = write_table(
        ws2, row2, "التغيرات (Delta)",
        ["المؤشر", "القيمة الحالية", "التغير"],
        delta_rows,
        title_color=BLUE
    )

    for r in range(row2 + 2, end_delta):
        cell = ws2.cell(r, 3)
        try:
            value = float(str(cell.value).replace("%", ""))
            if value > 0:
                cell.fill = fill(GREEN_SOFT)
                cell.font = font(10, True, "166534")
            elif value < 0:
                cell.fill = fill(RED_SOFT)
                cell.font = font(10, True, "991B1B")
            else:
                cell.fill = fill(SLATE_LIGHT)
                cell.font = font(10, True, SLATE)
        except Exception:
            pass

    FLAG_LABELS = {
        "water_low": "انخفاض الماء",
        "water_below_025": "ماء منخفض",
        "water_drop": "هبوط الماء",
        "siwsi_drop": "هبوط SIWSI",
        "ndre_low": "ضعف NDRE",
        "ndre_below_035": "NDRE منخفض",
        "ndvi_below_030": "NDVI منخفض",
        "ndvi_drop": "هبوط NDVI",
    }

    flags_rows = [[k, _safe_int(v)] for k, v in flag_counts.items()]
    flags_rows = sorted(flags_rows, key=lambda x: x[1], reverse=True)

    if not flags_rows:
        flags_rows = [["—", 0]]

    row2 = end_delta + 1
    end_flags = write_table(
        ws2, row2, "تفصيل إشارات الإجهاد",
        ["الإشارة", "العدد"],
        flags_rows,
        title_color=ORANGE
    )

    chart_flags_rows = [
        [FLAG_LABELS.get(k, k), v]
        for k, v in flags_rows
        if v > 0
    ][:6]

    if not chart_flags_rows:
        chart_flags_rows = [["لا توجد إشارات", 0]]

    chart_row2 = max(end_flags + 25, 80)
    ws2.cell(chart_row2, 1, "الإشارة")
    ws2.cell(chart_row2, 2, "العدد")

    for i, item in enumerate(chart_flags_rows, start=1):
        ws2.cell(chart_row2 + i, 1, item[0])
        ws2.cell(chart_row2 + i, 2, item[1])

    for r in range(chart_row2, chart_row2 + len(chart_flags_rows) + 2):
        ws2.row_dimensions[r].hidden = True

    flags_chart = BarChart()
    flags_chart.type = "bar"  
    flags_chart.style = 11
    flags_chart.title = "أكثر إشارات الإجهاد تكرارًا"
    flags_chart.height = 9.0
    flags_chart.width = 12.0
    flags_chart.legend = None
    flags_chart.gapWidth = 55

    flags_chart.add_data(
        Reference(ws2, min_col=2, min_row=chart_row2, max_row=chart_row2 + len(chart_flags_rows)),
        titles_from_data=True
    )
    flags_chart.set_categories(
        Reference(ws2, min_col=1, min_row=chart_row2 + 1, max_row=chart_row2 + len(chart_flags_rows))
    )

    flags_chart.dLbls = DataLabelList()
    flags_chart.dLbls.showVal = True
    flags_chart.dLbls.showLegendKey = False
    flags_chart.dLbls.showCatName = False
    flags_chart.dLbls.showSerName = False
    flags_chart.dLbls.showPercent = False

    try:
        flags_chart.dLbls.dLblPos = "outEnd"
    except Exception:
        pass

    flags_chart.x_axis.delete = False 
    flags_chart.y_axis.delete = False
    flags_chart.x_axis.title = None
    flags_chart.y_axis.title = None

    max_count = max([item[1] for item in chart_flags_rows] + [1])
    flags_chart.y_axis.scaling.min = 0
    flags_chart.y_axis.scaling.max = max_count * 1.25

    ws2.add_chart(flags_chart, f"D{row2}")

    # ─────────────────────────────────────────────
    # Sheet 3: المخاطر والمناطق
    # ─────────────────────────────────────────────
    ws3 = wb.create_sheet("المخاطر والمناطق")
    ws3.sheet_view.rightToLeft = True
    add_logo(ws3)
    for col, width in {
        "A": 22, "B": 14, "C": 14, "D": 46, "E": 10
    }.items():
        ws3.column_dimensions[col].width = width

    merge_value(
        ws3, "A1:E1", f"المخاطر والمناطق — {farm_name}",
        f=font(16, True, WHITE), bg=GREEN_DARK, al=align_center(), bd=border(GREEN_DARK)
    )

    risk_rows = []
    for item in risk_drivers[:10]:
        risk_rows.append([
            item.get("title", "—"),
            _safe_int(item.get("count", 0)),
            item.get("priority", "—"),
            item.get("note", "—"),
        ])
    if not risk_rows:
        risk_rows = [["—", 0, "—", "لا توجد بيانات متاحة"]]

    row3 = 3
    end_risk = write_table(
        ws3, row3, "أسباب الخطر الرئيسية",
        ["السبب", "عدد الإشارات", "الأولوية", "التفسير"],
        risk_rows,
        title_color=ORANGE
    )
    for r in range(row3 + 2, end_risk):
        cell = ws3.cell(r, 3)
        cell.fill = fill(priority_fill(cell.value))
        cell.font = font(10, True, "1E293B")

    hotspot_rows = []
    for i, item in enumerate(hotspots[:15], start=1):
        hotspot_rows.append([
            i,
            item.get("lat", "—"),
            item.get("lon", "—"),
            item.get("status", "—"),
            item.get("note", "—"),
        ])
    if not hotspot_rows:
        hotspot_rows = [["—", "—", "—", "—", "لا توجد نقاط بارزة حاليًا"]]

    row3 = end_risk + 1
    merge_value(
        ws3, f"A{row3}:E{row3}", "أهم المناطق التي تحتاج متابعة",
        f=font(12, True, WHITE), bg=RED, al=align_center(), bd=border()
    )
    headers = ["#", "خط العرض", "خط الطول", "الحالة", "الوصف"]
    for i, h in enumerate(headers, start=1):
        set_cell(
            ws3.cell(row3 + 1, i), h,
            f=font(10, True, SLATE), bg=SLATE_LIGHT, al=align_center(True), bd=border()
        )

    cur = row3 + 2
    for row_data in hotspot_rows:
        row_bg = status_fill(str(row_data[3]))
        for i, val in enumerate(row_data, start=1):
            set_cell(
                ws3.cell(cur, i), val,
                f=font(10, i == 4, "1E293B"),
                bg=row_bg,
                al=align_center(True) if i != 5 else align_right(True),
                bd=border()
            )
        cur += 1

    # risk chart data
    chart_row3 = max(cur + 25, 80)

    ws3.cell(chart_row3, 1, "السبب")
    ws3.cell(chart_row3, 2, "العدد")

    for i, item in enumerate(risk_rows, start=1):
        ws3.cell(chart_row3 + i, 1, item[0])
        ws3.cell(chart_row3 + i, 2, item[1])

    for r in range(chart_row3, chart_row3 + len(risk_rows) + 2):
        ws3.row_dimensions[r].hidden = True

    risk_chart = BarChart()
    risk_chart.title = "أسباب الخطر الرئيسية"
    risk_chart.style = 10
    risk_chart.height = 6.8
    risk_chart.width = 12

    risk_chart.add_data(
        Reference(ws3, min_col=2, min_row=chart_row3, max_row=chart_row3 + len(risk_rows)),
        titles_from_data=True
    )
    risk_chart.set_categories(
        Reference(ws3, min_col=1, min_row=chart_row3 + 1, max_row=chart_row3 + len(risk_rows))
    )

    style_chart_axes(
        risk_chart,
        y_title="عدد الإشارات",
        x_title="السبب",
        show_values=True,
        is_percent=False,
    )

    ws3.add_chart(risk_chart, f"A{cur + 1}")

    # ─────────────────────────────────────────────
    # Sheet 4: البيانات المكانية
    # ─────────────────────────────────────────────
    ws4 = wb.create_sheet("البيانات المكانية")
    ws4.sheet_view.rightToLeft = True
    add_logo(ws4)
    for col, width in {
        "A": 8, "B": 18, "C": 18, "D": 18, "E": 18
    }.items():
        ws4.column_dimensions[col].width = width

    merge_value(
        ws4, "A1:E1", f"البيانات المكانية — {farm_name}",
        f=font(16, True, WHITE), bg=GREEN_DARK, al=align_center(), bd=border(GREEN_DARK)
    )

    headers = ["م", "خط العرض", "خط الطول", "الحالة الحالية", "الحالة المتوقعة"]
    for i, h in enumerate(headers, start=1):
        set_cell(
            ws4.cell(3, i), h,
            f=font(10, True, SLATE), bg=SLATE_LIGHT, al=align_center(True), bd=border()
        )

    r = 4
    for idx, pt in enumerate(map_points, start=1):
        s_text = decode_status(pt.get("currentStatus"))
        ps_text = decode_status(pt.get("predictedStatus"))
        row_bg = status_fill(s_text)
        values = [idx, pt.get("lat", "—"), pt.get("lng", "—"), s_text, ps_text]
        for c, val in enumerate(values, start=1):
            set_cell(
                ws4.cell(r, c), val,
                f=font(10, c in [4, 5], "1E293B"),
                bg=row_bg,
                al=align_center(True),
                bd=border()
            )
        r += 1

    # ─────────────────────────────────────────────
    # Sheet 5: المسار الزمني
    # ─────────────────────────────────────────────
    ws5 = wb.create_sheet("المسار الزمني")
    ws5.sheet_view.rightToLeft = True
    add_logo(ws5)
    for col, width in {"A": 18, "B": 14, "C": 14, "D": 14}.items():
        ws5.column_dimensions[col].width = width

    merge_value(
        ws5, "A1:H1", f"المسار الزمني — {farm_name}",
        f=font(16, True, WHITE), bg=GREEN_DARK, al=align_center(), bd=border(GREEN_DARK)
    )

    trend_headers = ["التاريخ", "NDVI", "NDMI", "NDRE"]
    for i, h in enumerate(trend_headers, start=1):
        set_cell(
            ws5.cell(3, i), h,
            f=font(10, True, SLATE), bg=SLATE_LIGHT, al=align_center(), bd=border()
        )

    trend_dates = safe_list(multi_trend.get("dates"))
    trend_ndvi = safe_list(multi_trend.get("ndvi"))
    trend_ndmi = safe_list(multi_trend.get("ndmi"))
    trend_ndre = safe_list(multi_trend.get("ndre"))
    max_len = max(len(trend_dates), len(trend_ndvi), len(trend_ndmi), len(trend_ndre), 0)

    for i in range(max_len):
        vals = [
            trend_dates[i] if i < len(trend_dates) else "—",
            _safe_float(trend_ndvi[i]) if i < len(trend_ndvi) else None,
            _safe_float(trend_ndmi[i]) if i < len(trend_ndmi) else None,
            _safe_float(trend_ndre[i]) if i < len(trend_ndre) else None,
        ]
        for j, val in enumerate(vals, start=1):
            set_cell(
                ws5.cell(i + 4, j), val,
                f=font(10, False, "1E293B"),
                al=align_center(),
                bd=border()
            )

    if max_len > 0:
        lc = LineChart()
        lc.title = "تغير المؤشرات عبر الزمن"
        lc.style = 10
        lc.height = 10
        lc.width = 16
        lc.add_data(
            Reference(ws5, min_col=2, max_col=4, min_row=3, max_row=3 + max_len),
            titles_from_data=True
        )
        lc.set_categories(
            Reference(ws5, min_col=1, min_row=4, max_row=3 + max_len)
        )
        style_chart_axes(
            lc,
            y_title="القيمة",
            x_title="التاريخ",
            show_values=False,
            is_percent=False,
        )
        ws5.add_chart(lc, "F3")

    # ─────────────────────────────────────────────
    # Sheet 6: مصدر الرسوم فقط
    # ─────────────────────────────────────────────
    ws6 = wb.create_sheet("chart_data")
    ws6.sheet_state = "hidden"

    # page1
    ws6["A1"] = "الفئة"
    ws6["B1"] = "الحالي"
    ws6["C1"] = "المتوقع"
    ws6["A2"] = "سليم";   ws6["B2"] = healthy_pct;   ws6["C2"] = healthy_next
    ws6["A3"] = "متابعة"; ws6["B3"] = monitor_pct;   ws6["C3"] = monitor_next
    ws6["A4"] = "حرج";    ws6["B4"] = critical_pct;  ws6["C4"] = critical_next

    # page2
    ws6["E1"] = "الإشارة"
    ws6["F1"] = "العدد"
    for i, item in enumerate(chart_flags_rows, start=2):
        ws6.cell(i, 5, item[0])
        ws6.cell(i, 6, item[1])

    # page3
    ws6["H1"] = "السبب"
    ws6["I1"] = "العدد"
    for i, item in enumerate(risk_rows, start=2):
        ws6.cell(i, 8, item[0])
        ws6.cell(i, 9, item[1])

    # ─────────────────────────────────────────────
    #  الرسوم
    # ─────────────────────────────────────────────
    # pie
    pie = PieChart()
    pie.title = "التوزيع الحالي"
    pie.height = 7.8
    pie.width = 10.2

    pie.add_data(
        Reference(ws6, min_col=2, min_row=1, max_row=4),
        titles_from_data=True
    )
    pie.set_categories(
        Reference(ws6, min_col=1, min_row=2, max_row=4)
    )

    pie_colors = ["22C55E", "2563EB", "EF4444"]

    try:
        pie.series[0].data_points = []
        for idx, color in enumerate(pie_colors):
            dp = DataPoint(idx=idx)
            dp.graphicalProperties.solidFill = color
            pie.series[0].data_points.append(dp)
    except Exception:
        pass

    pie.dLbls = DataLabelList()
    pie.dLbls.showVal = True
    pie.dLbls.showPercent = True
    pie.dLbls.showCatName = True
    pie.dLbls.showLeaderLines = True

    ws.add_chart(pie, "A37")

    # bar
    bar.series = []
    bar.add_data(
        Reference(ws6, min_col=2, max_col=3, min_row=1, max_row=4),
        titles_from_data=True
    )
    bar.set_categories(
        Reference(ws6, min_col=1, min_row=2, max_row=4)
    )

    # flags chart
    flags_chart.series = []
    flags_chart.add_data(
        Reference(ws6, min_col=6, min_row=1, max_row=1 + len(chart_flags_rows)),
        titles_from_data=True
    )
    flags_chart.set_categories(
        Reference(ws6, min_col=5, min_row=2, max_row=1 + len(chart_flags_rows))
    )

    flags_chart.dLbls = DataLabelList()
    flags_chart.dLbls.showVal = True
    flags_chart.dLbls.showLegendKey = False
    flags_chart.dLbls.showCatName = False
    flags_chart.dLbls.showSerName = False
    flags_chart.dLbls.showPercent = False

    try:
        flags_chart.dLbls.dLblPos = "outEnd"
    except Exception:
        pass

    flags_chart.legend = None
    flags_chart.x_axis.title = None
    flags_chart.y_axis.title = None
    flags_chart.x_axis.scaling.min = 0
    flags_chart.x_axis.scaling.max = max_count * 1.25

    # risk chart
    risk_chart.series = []
    risk_chart.add_data(
        Reference(ws6, min_col=9, min_row=1, max_row=1 + len(risk_rows)),
        titles_from_data=True
    )
    risk_chart.set_categories(
        Reference(ws6, min_col=8, min_row=2, max_row=1 + len(risk_rows))
    )

    # ─────────────────────────────────────────────
    # final formatting
    # ─────────────────────────────────────────────
    for sheet in wb.worksheets:
        sheet.sheet_view.showGridLines = False

    output_path = f"/tmp/saaf_report_{farm_id}.xlsx"
    wb.save(output_path)
    return output_path

def _first_non_empty(*values):
    for v in values:
        if v is None:
            continue
        if isinstance(v, str) and not v.strip():
            continue
        if isinstance(v, (list, dict, tuple, set)) and len(v) == 0:
            continue
        if v == "":
            continue
        return v
    return None

def _prefer_live_number(old_value, live_value, default=0):
    """
    إذا القيمة القديمة None/فارغة نأخذ الحية.
    وإذا القديمة = 0 لكن الحية غير صفر، نأخذ الحية.
    وإلا نُبقي القديمة.
    """
    try:
        if old_value is None:
            return live_value if live_value is not None else default

        if isinstance(old_value, str) and not old_value.strip():
            return live_value if live_value is not None else default

        old_num = float(old_value)
        live_num = float(live_value) if live_value is not None else None

        if (math.isnan(old_num) or old_num == 0) and live_num not in (None, 0):

            return live_value

        return old_value
    except Exception:
        return live_value if live_value is not None else default

def _farm_centroid_from_polygon(poly: list | None):
    pts = _normalize_polygon(poly)
    if len(pts) < 3:
        return None, None
    lat = sum(p[0] for p in pts) / len(pts)
    lng = sum(p[1] for p in pts) / len(pts)
    return lat, lng

def _get_report_weather_live(farm_data: dict) -> dict:
    if not WEATHERAPI_KEY:
        logger.error("WEATHERAPI_KEY missing in reports route")
        return {"rain_mm": 0.0, "t_mean": 0.0}

    poly = farm_data.get("polygon") or []
    lat, lng = _farm_centroid_from_polygon(poly)

    if lat is None or lng is None:
        logger.error("Farm polygon missing/invalid for live weather")
        return {"rain_mm": 0.0, "t_mean": 0.0}

    logger.info("Weather debug | centroid lat=%s lng=%s", lat, lng)

    try:
        end_dt = datetime.utcnow().date()
        start_dt = end_dt - __import__("datetime").timedelta(days=7)

        rows = []
        current = start_dt

        while current <= end_dt:
            d = current.strftime("%Y-%m-%d")
            url = "https://api.weatherapi.com/v1/history.json"
            params = {
                "key": WEATHERAPI_KEY,
                "q": f"{lat},{lng}",
                "dt": d,
            }

            resp = requests.get(url, params=params, timeout=20)
            logger.info("WeatherAPI request | date=%s | status=%s", d, resp.status_code)

            if resp.status_code != 200:
                logger.error("WeatherAPI failed | date=%s | body=%s", d, resp.text[:500])
                resp.raise_for_status()

            data = resp.json()
            day = ((data.get("forecast") or {}).get("forecastday") or [{}])[0].get("day", {})

            rows.append({
                "precip_mm": float(day.get("totalprecip_mm", 0) or 0),
                "t2m_mean": float(day.get("avgtemp_c", 0) or 0),
            })

            current += __import__("datetime").timedelta(days=1)

        if not rows:
            logger.error("WeatherAPI returned no rows")
            return {"rain_mm": 0.0, "t_mean": 0.0}

        rain_mm = round(sum(r["precip_mm"] for r in rows), 1)
        t_mean = round(sum(r["t2m_mean"] for r in rows) / len(rows), 1)

        logger.info("Live WeatherAPI success | rain_mm=%s | t_mean=%s", rain_mm, t_mean)
        return {"rain_mm": rain_mm, "t_mean": t_mean}

    except Exception:
        logger.error("Live WeatherAPI crashed:\n%s", traceback.format_exc())
        return {"rain_mm": 0.0, "t_mean": 0.0}

def _merge_export_with_live_farm_data(export_data: dict, farm_data: dict) -> dict:
    export_data = dict(export_data or {})
    farm_data = farm_data or {}

    health_root = farm_data.get("health") if isinstance(farm_data.get("health"), dict) else {}
    current_health = health_root.get("current_health") if isinstance(health_root.get("current_health"), dict) else {}
    forecast_next = health_root.get("forecast_next_week") if isinstance(health_root.get("forecast_next_week"), dict) else {}

    header = dict(export_data.get("header", {}) or {})
    header["name"] = _first_non_empty(header.get("name"), farm_data.get("farmName"), "—")
    header["area"] = _first_non_empty(header.get("area"), farm_data.get("farmSize"), "—")
    header["city"] = _first_non_empty(header.get("city"), farm_data.get("region"), farm_data.get("city"), "—")
    header["contract_number"] = _first_non_empty(header.get("contract_number"), farm_data.get("contractNumber"), "—")
    header["owner_name"] = _first_non_empty(header.get("owner_name"), farm_data.get("ownerName"), "—")
    header["total_palms"] = _first_non_empty(
        header.get("total_palms"),
        export_data.get("finalCount"),
        farm_data.get("finalCount"),
        farm_data.get("palmCount"),
        farm_data.get("totalPalms"),
        0,
    )
    export_data["header"] = header

    export_data["farmName"] = _first_non_empty(export_data.get("farmName"), farm_data.get("farmName"))
    export_data["farmSize"] = _first_non_empty(export_data.get("farmSize"), farm_data.get("farmSize"))
    export_data["finalCount"] = _first_non_empty(
        export_data.get("finalCount"),
        farm_data.get("finalCount"),
        farm_data.get("palmCount"),
        farm_data.get("totalPalms"),
        0,
    )
    export_data["owner_name"] = _first_non_empty(export_data.get("owner_name"), farm_data.get("ownerName"), "—")

    export_data["farm_polygon"] = _first_non_empty(
        export_data.get("farm_polygon"),
        farm_data.get("polygon"),
        [],
    )

    export_data["health_map_points"] = _first_non_empty(
        export_data.get("health_map_points"),
        farm_data.get("healthMap"),
        health_root.get("health_map"),
        [],
    )

    dist = dict(export_data.get("distribution", {}) or {})
    export_data["distribution"] = {
        "Healthy_Pct": _prefer_live_number(dist.get("Healthy_Pct"), current_health.get("Healthy_Pct"), 0),
        "Monitor_Pct": _prefer_live_number(dist.get("Monitor_Pct"), current_health.get("Monitor_Pct"), 0),
        "Critical_Pct": _prefer_live_number(dist.get("Critical_Pct"), current_health.get("Critical_Pct"), 0),
    }

    next_week = dict(export_data.get("forecast_next_week", {}) or {})
    export_data["forecast_next_week"] = {
        "Healthy_Pct_next": _prefer_live_number(next_week.get("Healthy_Pct_next"), forecast_next.get("Healthy_Pct_next"), 0),
        "Monitor_Pct_next": _prefer_live_number(next_week.get("Monitor_Pct_next"), forecast_next.get("Monitor_Pct_next"), 0),
        "Critical_Pct_next": _prefer_live_number(next_week.get("Critical_Pct_next"), forecast_next.get("Critical_Pct_next"), 0),
        "ndvi_delta_next_mean": _prefer_live_number(next_week.get("ndvi_delta_next_mean"), forecast_next.get("ndvi_delta_next_mean"), 0),
        "ndmi_delta_next_mean": _prefer_live_number(next_week.get("ndmi_delta_next_mean"), forecast_next.get("ndmi_delta_next_mean"), 0),
    }

    biometrics = dict(export_data.get("biometrics", {}) or {})
    export_data["biometrics"] = {
        "ndvi": {
            "val": _prefer_live_number(
                (biometrics.get("ndvi") or {}).get("val") if isinstance(biometrics.get("ndvi"), dict) else None,
                _first_non_empty(current_health.get("NDVI"), current_health.get("ndvi_mean")),
                0,
            ),
            "delta": _prefer_live_number(
                (biometrics.get("ndvi") or {}).get("delta") if isinstance(biometrics.get("ndvi"), dict) else None,
                current_health.get("NDVI_delta"),
                0,
            ),
        },
        "ndmi": {
            "val": _prefer_live_number(
                (biometrics.get("ndmi") or {}).get("val") if isinstance(biometrics.get("ndmi"), dict) else None,
                _first_non_empty(current_health.get("NDMI"), current_health.get("ndmi_mean")),
                0,
            ),
            "delta": _prefer_live_number(
                (biometrics.get("ndmi") or {}).get("delta") if isinstance(biometrics.get("ndmi"), dict) else None,
                current_health.get("NDMI_delta"),
                0,
            ),
        },
        "ndre": {
            "val": _prefer_live_number(
                (biometrics.get("ndre") or {}).get("val") if isinstance(biometrics.get("ndre"), dict) else None,
                _first_non_empty(current_health.get("NDRE"), current_health.get("ndre_mean")),
                0,
            ),
            "delta": _prefer_live_number(
                (biometrics.get("ndre") or {}).get("delta") if isinstance(biometrics.get("ndre"), dict) else None,
                current_health.get("NDRE_delta"),
                0,
            ),
        },
    }

    climate = dict(export_data.get("climate", {}) or {})

    need_live_weather = (
    _safe_float(climate.get("rain_mm"), 0.0) == 0.0 or
    _safe_float(climate.get("t_mean"), 0.0) == 0.0
     )

    live_weather = _get_report_weather_live(farm_data) if need_live_weather else {}

    export_data["climate"] = {
        **climate,
        "rain_mm": _prefer_live_number(climate.get("rain_mm"), live_weather.get("rain_mm"), 0.0),
        "t_mean": _prefer_live_number(climate.get("t_mean"), live_weather.get("t_mean"), 0.0),
        "total_pixels": _prefer_live_number(climate.get("total_pixels"), current_health.get("total_pixels"), 0),
        "rpw_score": _prefer_live_number(climate.get("rpw_score"), current_health.get("rpw_score"), 0),
    }

    risk_diagnostics = (
        health_root.get("risk_diagnostics")
        or health_root.get("alert_signals")
        or {}
    )

    alert_context = dict(export_data.get("alert_context", {}) or {})
    export_data["alert_context"] = {
        **alert_context,
        "total_pixels": _prefer_live_number(
            alert_context.get("total_pixels"),
            current_health.get("total_pixels"),
            0,
        ),
        "pixels_with_any_flag": _prefer_live_number(
            alert_context.get("pixels_with_any_flag"),
            risk_diagnostics.get("affected_pixels_count")
            or risk_diagnostics.get("pixels_with_any_flag_latest"),
            0,
        ),
        "flag_counts": _first_non_empty(alert_context.get("flag_counts"), {}),
    }

    if not export_data.get("multi_trend"):
        hist = health_root.get("indices_history_last_month", []) or []
        export_data["multi_trend"] = {
            "dates": [x.get("date") for x in hist if isinstance(x, dict)],
            "ndvi": [x.get("NDVI") for x in hist if isinstance(x, dict)],
            "ndmi": [x.get("NDMI") for x in hist if isinstance(x, dict)],
            "ndre": [x.get("NDRE") for x in hist if isinstance(x, dict)],
        }

    return export_data
# ─────────────────────────────────────────────
# Routes
# ─────────────────────────────────────────────
def _deg_to_tile(lat: float, lon: float, zoom: int):
    lat_rad = math.radians(lat)
    n = 2.0 ** zoom
    xtile = int((lon + 180.0) / 360.0 * n)
    ytile = int(
        n * (1.0 - math.log(math.tan(lat_rad) + (1 / math.cos(lat_rad))) / math.pi) / 2.0
    )
    return xtile, ytile

def _latlon_to_global_pixels(lat: float, lon: float, zoom: int):
    siny = math.sin(math.radians(lat))
    siny = min(max(siny, -0.9999), 0.9999)

    scale = TILE_SIZE_MAP * (2 ** zoom)
    x = (lon + 180.0) / 360.0 * scale
    y = (0.5 - math.log((1 + siny) / (1 - siny)) / (4 * math.pi)) * scale
    return x, y

def _stitch_maptiler_tiles(bounds: dict, width: int, height: int, zoom: int | None = None) -> dict:
    try:
        api_key = os.environ.get("MAPTILER_KEY", "").strip()
        if not api_key or not bounds:
            return {"bg_data_uri": None, "meta": None}

        pad = 8

        min_lat = bounds["min_lat"]
        max_lat = bounds["max_lat"]
        min_lng = bounds["min_lng"]
        max_lng = bounds["max_lng"]

        # نختار زوم يخلي كامل حدود المزرعة تدخل داخل الصورة
        chosen_zoom = 16
        for z in range(18, 9, -1):
            gx1, gy1 = _latlon_to_global_pixels(max_lat, min_lng, z)  # top-left
            gx2, gy2 = _latlon_to_global_pixels(min_lat, max_lng, z)  # bottom-right

            req_w = abs(gx2 - gx1)
            req_h = abs(gy2 - gy1)

            if req_w <= (width - pad * 2) and req_h <= (height - pad * 2):
                chosen_zoom = z
                break

        if zoom is not None:
            chosen_zoom = zoom

        gx1, gy1 = _latlon_to_global_pixels(max_lat, min_lng, chosen_zoom)  # top-left
        gx2, gy2 = _latlon_to_global_pixels(min_lat, max_lng, chosen_zoom)  # bottom-right

        farm_w = abs(gx2 - gx1)
        farm_h = abs(gy2 - gy1)

        # نخلي حدود المزرعة بمنتصف الكانفس
        origin_x = min(gx1, gx2) - ((width - farm_w) / 2.0)
        origin_y = min(gy1, gy2) - ((height - farm_h) / 2.0)

        end_x = origin_x + width
        end_y = origin_y + height

        start_tile_x = int(math.floor(origin_x / TILE_SIZE_MAP))
        start_tile_y = int(math.floor(origin_y / TILE_SIZE_MAP))
        end_tile_x = int(math.floor(end_x / TILE_SIZE_MAP))
        end_tile_y = int(math.floor(end_y / TILE_SIZE_MAP))

        stitched_w = (end_tile_x - start_tile_x + 1) * TILE_SIZE_MAP
        stitched_h = (end_tile_y - start_tile_y + 1) * TILE_SIZE_MAP
        stitched = Image.new("RGB", (stitched_w, stitched_h))

        for tx in range(start_tile_x, end_tile_x + 1):
            for ty in range(start_tile_y, end_tile_y + 1):
                url = REPORT_TILE_URL.format(zoom=chosen_zoom, x=tx, y=ty, key=api_key)

                resp = requests.get(url, timeout=20)
                resp.raise_for_status()

                tile_img = Image.open(io.BytesIO(resp.content)).convert("RGB")
                px = (tx - start_tile_x) * TILE_SIZE_MAP
                py = (ty - start_tile_y) * TILE_SIZE_MAP
                stitched.paste(tile_img, (px, py))

        crop_x = int(round(origin_x - start_tile_x * TILE_SIZE_MAP))
        crop_y = int(round(origin_y - start_tile_y * TILE_SIZE_MAP))

        stitched = stitched.crop((crop_x, crop_y, crop_x + width, crop_y + height))

        buffer = io.BytesIO()
        stitched.save(buffer, format="JPEG", quality=85)
        img_b64 = base64.b64encode(buffer.getvalue()).decode("utf-8")

        meta = {
            "zoom": chosen_zoom,
            "origin_x": origin_x,
            "origin_y": origin_y,
        }

        return {
            "bg_data_uri": f"data:image/jpeg;base64,{img_b64}",
            "meta": meta,
        }

    except Exception as e:
        logger.warning(f"Failed to stitch MapTiler tiles: {e}")
        return {"bg_data_uri": None, "meta": None}
    
@reports_bp.route('/reports/<farm_id>/pdf', methods=['GET'])
def export_pdf(farm_id):
    try:
        doc, _ = get_farm_safely(farm_id)
        if not doc:
            return jsonify({"ok": False, "error": "المزرعة غير موجودة"}), 404

        farm_data = doc.to_dict() or {}

        if not farm_data.get("health"):
            return jsonify({"ok": False, "error": "بيانات التحليل ناقصة. شغّل التحليل أولاً."}), 400

        stored_export = farm_data.get("export_data") or {}

        if stored_export:
            export_data = _merge_export_with_live_farm_data(stored_export, farm_data)
        else:
            export_data = prepare_export_data(
                farm_data,
                farm_data["health"],
                detected_count=int(
                    farm_data.get("palm_count")
                    or farm_data.get("finalCount")
                    or 0
                ),
                        )
            export_data = _merge_export_with_live_farm_data(export_data, farm_data)

        logger.info(
            "PDF export debug | farm_id=%s | healthMap=%s | export health_map_points=%s | total_pixels=%s | rain_mm=%s | t_mean=%s",
            farm_id,
            len(farm_data.get("healthMap", []) or []),
            len(export_data.get("health_map_points", []) or []),
            ((export_data.get("alert_context", {}) or {}).get("total_pixels")),
            ((export_data.get("climate", {}) or {}).get("rain_mm")),
            ((export_data.get("climate", {}) or {}).get("t_mean")),
        )

        pdf_path = generate_pdf_report(export_data, farm_id, farm_doc=farm_data)
        with open(pdf_path, "rb") as f:
            encoded = base64.b64encode(f.read()).decode('utf-8')

        return jsonify({
            "ok": True,
            "pdfBase64": encoded,
            "fileName": f"{farm_data.get('farmName', 'Farm')}_farm_report.pdf"
        }), 200

    except Exception as e:
        logger.error(f"💥 PDF Route Crash: {traceback.format_exc()}")
        return jsonify({
            "ok": False,
            "error": str(e),
            "trace": traceback.format_exc()
        }), 500


@reports_bp.route('/reports/<farm_id>/excel', methods=['GET'])
def export_excel(farm_id):
    try:
        doc, _ = get_farm_safely(farm_id)
        if not doc:
            return jsonify({"ok": False, "error": "المزرعة غير موجودة"}), 404

        farm_data = doc.to_dict() or {}

        if not farm_data.get("health"):
            return jsonify({"ok": False, "error": "بيانات التحليل ناقصة. شغّل التحليل أولاً."}), 400

        stored_export = farm_data.get("export_data") or {}

        if stored_export:
            export_data = _merge_export_with_live_farm_data(stored_export, farm_data)
        else:
            export_data = prepare_export_data(
                farm_data,
                farm_data["health"],
                detected_count=int(
                    farm_data.get("palm_count")
                    or farm_data.get("finalCount")
                    or 0
                ),
            )
            export_data = _merge_export_with_live_farm_data(export_data, farm_data)

        excel_path = generate_excel_report(export_data, farm_id)
        with open(excel_path, "rb") as f:
            encoded = base64.b64encode(f.read()).decode('utf-8')

        return jsonify({
            "ok": True,
            "excelBase64": encoded,
            "fileName": f"{farm_data.get('farmName', 'Farm')}_farm_report.xlsx"
        }), 200

    except Exception as e:
        logger.error(f"💥 Excel Route Crash: {traceback.format_exc()}")
        return jsonify({"ok": False, "error": str(e), "trace": traceback.format_exc()}), 500